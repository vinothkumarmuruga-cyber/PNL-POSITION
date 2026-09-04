"""
Upstox Live LTP — Option Strategy P&L tracker (Streamlit)
-----------------------------------------------------------
Reads your Option_Strategy_PNL.xlsx layout (Monthly - D / Monthly -ND /
Weekly - D, 2 rows per position = CE leg + PE leg), lets you fill in each
leg's expiry, resolves Upstox instrument keys, pulls live LTP via the
Upstox v3 Market Quote API, and shows live points / live P&L per position.

Run locally:
    pip install -r requirements.txt
    streamlit run app.py

Deploy on Streamlit Cloud:
    Push this folder to a repo, deploy on share.streamlit.io, then in
    App settings -> Secrets paste:
        UPSTOX_CLIENT_ID = "..."
        UPSTOX_CLIENT_SECRET = "..."
        REDIRECT_URI = "https://<your-app>.streamlit.app"
    Register that same REDIRECT_URI in your Upstox developer app.
"""

import gzip
import io
import re
from datetime import date, datetime

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook

from positions_data import POSITIONS  # bundled snapshot; re-export after editing (see sidebar)

try:
    from streamlit_autorefresh import st_autorefresh
    HAS_AUTOREFRESH = True
except ImportError:
    HAS_AUTOREFRESH = False

st.set_page_config(page_title="Option Strategy P&L — Live LTP", layout="wide")

SHEET_NAMES = ['Monthly - D ', 'Monthly -ND', 'Weekly - D ']  # exact tab names, incl. trailing spaces
AUTHORIZE_URL = "https://api.upstox.com/v2/login/authorization/dialog"
TOKEN_URL = "https://api.upstox.com/v2/login/authorization/token"
LTP_URL = "https://api.upstox.com/v3/market-quote/ltp"
INSTRUMENT_MASTER_URL = "https://assets.upstox.com/market-quote/instruments/exchange/NSE.json.gz"

# ---------------------------------------------------------------- helpers

def normalize_symbol(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def parse_strike_cell(text):
    m = re.match(r"^(\d+(?:\.\d+)?)\s*(CE|PE)$", str(text or "").strip(), re.IGNORECASE)
    if not m:
        return None
    return float(m.group(1)), m.group(2).upper()


def cell_or_none(v):
    return None if v in ("", None) else v


def positions_to_df(positions: list) -> pd.DataFrame:
    df = pd.DataFrame(positions)
    for col in ("entry_date", "exit_date"):
        df[col] = pd.to_datetime(df[col], errors="coerce")
    df["expiry_date"] = pd.to_datetime(df["expiry_date"], errors="coerce").dt.date
    df["instrument_key"] = df["instrument_key"].fillna("")
    return df


# ---------------------------------------------------------------- parse workbook -> positions

def parse_workbook(file) -> pd.DataFrame:
    """Returns one row per leg: sheet, position_no, leg (1/2), symbol, strike_text,
    strike, opt_type, lot_size, qty, entry, exit, entry_date, exit_date, remarks."""
    wb = load_workbook(file, data_only=False)
    rows = []
    for sheet_name in SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=2).value == "S.no":
                header_row = r
                break
        if header_row is None:
            continue

        r = header_row + 1
        max_r = ws.max_row
        while r <= max_r:
            sno = ws.cell(row=r, column=2).value
            if isinstance(sno, (int, float)):
                leg1_row = r
                leg2_row = r + 1
                leg2_f = ws.cell(row=leg2_row, column=6).value if leg2_row <= max_r else None
                leg2_b = ws.cell(row=leg2_row, column=2).value if leg2_row <= max_r else None
                has_leg2 = bool(leg2_f) and not isinstance(leg2_b, (int, float))

                lot_size = ws.cell(row=leg1_row, column=7).value
                entry_date = ws.cell(row=leg1_row, column=3).value
                strategy = ws.cell(row=leg1_row, column=4).value
                symbol = ws.cell(row=leg1_row, column=5).value

                for leg_num, leg_row in ([(1, leg1_row), (2, leg2_row)] if has_leg2 else [(1, leg1_row)]):
                    strike_text = ws.cell(row=leg_row, column=6).value
                    parsed = parse_strike_cell(strike_text)
                    rows.append({
                        "sheet": sheet_name.strip(),
                        "position_no": int(sno),
                        "row": leg_row,
                        "leg": leg_num,
                        "entry_date": entry_date,
                        "strategy": strategy,
                        "symbol": symbol,
                        "strike_text": strike_text,
                        "strike": parsed[0] if parsed else None,
                        "opt_type": parsed[1] if parsed else None,
                        "lot_size": lot_size,
                        "qty": cell_or_none(ws.cell(row=leg_row, column=8).value),
                        "entry": cell_or_none(ws.cell(row=leg_row, column=9).value),
                        "exit": cell_or_none(ws.cell(row=leg_row, column=10).value),
                        "exit_date": ws.cell(row=leg_row, column=17).value,
                        "remarks": ws.cell(row=leg_row, column=18).value,
                        "expiry_date": None,
                        "instrument_key": "",
                    })
                r = (leg2_row if has_leg2 else leg1_row) + 1
                continue
            r += 1
    return pd.DataFrame(rows)


# ---------------------------------------------------------------- Upstox: instrument master

@st.cache_data(ttl=6 * 3600, show_spinner="Downloading & filtering Upstox instrument master...")
def fetch_instrument_master(symbols_key: tuple) -> pd.DataFrame:
    symbols = set(symbols_key)
    resp = requests.get(INSTRUMENT_MASTER_URL, timeout=120)
    resp.raise_for_status()
    data = gzip.decompress(resp.content)
    import json
    all_records = json.loads(data)
    kept = []
    for rec in all_records:
        if rec.get("segment") != "NSE_FO":
            continue
        if rec.get("instrument_type") not in ("CE", "PE"):
            continue
        norm = normalize_symbol(rec.get("name"))
        if norm not in symbols:
            continue
        kept.append({
            "name": rec.get("name"),
            "trading_symbol": rec.get("trading_symbol"),
            "instrument_key": rec.get("instrument_key"),
            "strike": rec.get("strike_price"),
            "type": rec.get("instrument_type"),
            "expiry_ms": rec.get("expiry"),
        })
    return pd.DataFrame(kept)


def match_instrument_key(symbol, strike, opt_type, expiry: date, master: pd.DataFrame):
    if master.empty or not expiry or strike is None or opt_type is None:
        return None, "no data"
    norm_symbol = normalize_symbol(symbol)
    cand = master[(master["strike"] == strike) & (master["type"] == opt_type)]
    cand = cand[cand["expiry_ms"].apply(
        lambda ms: bool(ms) and datetime.fromtimestamp(ms / 1000).date() == expiry
    )]
    if cand.empty:
        return None, "NOT FOUND"
    exact = cand[cand["name"].apply(normalize_symbol) == norm_symbol]
    pool = exact if not exact.empty else cand[cand["name"].apply(
        lambda n: normalize_symbol(n) in norm_symbol or norm_symbol in normalize_symbol(n)
    )]
    if len(pool) == 1:
        return pool.iloc[0]["instrument_key"], "matched"
    if len(pool) == 0:
        return None, "NOT FOUND"
    return None, f"AMBIGUOUS ({len(pool)})"


# ---------------------------------------------------------------- Upstox: OAuth + LTP

def authorize_url():
    client_id = st.secrets.get("UPSTOX_CLIENT_ID", "")
    redirect_uri = st.secrets.get("REDIRECT_URI", "")
    return (f"{AUTHORIZE_URL}?client_id={client_id}&redirect_uri={redirect_uri}"
            f"&response_type=code")


def exchange_code_for_token(code: str):
    resp = requests.post(TOKEN_URL, data={
        "code": code,
        "client_id": st.secrets.get("UPSTOX_CLIENT_ID", ""),
        "client_secret": st.secrets.get("UPSTOX_CLIENT_SECRET", ""),
        "redirect_uri": st.secrets.get("REDIRECT_URI", ""),
        "grant_type": "authorization_code",
    }, headers={"accept": "application/json"}, timeout=30)
    resp.raise_for_status()
    return resp.json()["access_token"]


def fetch_ltp(instrument_keys: list, access_token: str) -> dict:
    """Returns {instrument_key: last_price}."""
    out = {}
    for i in range(0, len(instrument_keys), 500):
        batch = instrument_keys[i:i + 500]
        resp = requests.get(LTP_URL, params={"instrument_key": ",".join(batch)},
                             headers={"Authorization": f"Bearer {access_token}", "Accept": "application/json"},
                             timeout=30)
        if resp.status_code == 401:
            raise RuntimeError("Upstox token expired/invalid — reconnect.")
        resp.raise_for_status()
        for rec in resp.json().get("data", {}).values():
            tok = rec.get("instrument_token")
            if tok:
                out[tok] = rec.get("last_price")
    return out


# ---------------------------------------------------------------- app state

if "df" not in st.session_state:
    st.session_state.df = positions_to_df(POSITIONS)  # loads bundled positions_data.py on first run
if "access_token" not in st.session_state:
    st.session_state.access_token = None

# handle OAuth redirect
qp = st.query_params
if "code" in qp and not st.session_state.access_token:
    try:
        st.session_state.access_token = exchange_code_for_token(qp["code"])
        st.query_params.clear()
        st.success("Connected to Upstox.")
    except Exception as e:
        st.error(f"Token exchange failed: {e}")

# ---------------------------------------------------------------- sidebar

st.sidebar.header("Upstox connection")
if st.session_state.access_token:
    st.sidebar.success("Connected. Reconnect daily (token expires 3:30 AM IST).")
else:
    st.sidebar.warning("Not connected.")
    st.sidebar.link_button("Login to Upstox", authorize_url())

st.sidebar.divider()
st.sidebar.caption(f"Loaded from bundled positions_data.py ({len(POSITIONS)} legs) unless you upload a newer workbook below.")
uploaded = st.sidebar.file_uploader("Upload a newer Option_Strategy_PNL.xlsx", type=["xlsx"])
if uploaded is not None and st.sidebar.button("Load uploaded workbook"):
    st.session_state.df = parse_workbook(uploaded)
    st.sidebar.success(f"Loaded {len(st.session_state.df)} legs.")
if st.sidebar.button("Reset to bundled positions_data.py"):
    st.session_state.df = positions_to_df(POSITIONS)
    st.rerun()

if HAS_AUTOREFRESH:
    refresh_on = st.sidebar.checkbox("Auto-refresh LTP every 30s", value=False)
    if refresh_on:
        st_autorefresh(interval=30_000, key="ltp_autorefresh")

# ---------------------------------------------------------------- main

st.title("Option Strategy P&L — Live LTP")

df = st.session_state.df

st.subheader("1. Fill Expiry Date for open legs, sync & match instrument keys")
edit_cols = ["sheet", "position_no", "leg", "symbol", "strike_text", "entry", "exit", "expiry_date", "instrument_key"]
edited = st.data_editor(
    df[edit_cols],
    column_config={
        "expiry_date": st.column_config.DateColumn("Expiry Date"),
        "instrument_key": st.column_config.TextColumn("Instrument Key (auto or paste manually)"),
    },
    disabled=["sheet", "position_no", "leg", "symbol", "strike_text", "entry", "exit"],
    width="stretch", hide_index=True, key="editor",
)
df["expiry_date"] = edited["expiry_date"]
df["instrument_key"] = edited["instrument_key"]

col_a, col_b, col_c = st.columns(3)

if col_a.button("Sync Instrument Master"):
    symbols = tuple(sorted(set(normalize_symbol(s) for s in df["symbol"].dropna())))
    st.session_state.master = fetch_instrument_master(symbols)
    st.success(f"Cached {len(st.session_state.master)} contracts.")

if col_b.button("Match Instrument Keys"):
    master = st.session_state.get("master")
    if master is None or master.empty:
        st.error("Sync instrument master first.")
    else:
        matched = notfound = ambiguous = 0
        for idx, row in df.iterrows():
            if row["exit"] not in (None, "") or row["instrument_key"]:
                continue
            key, status = match_instrument_key(row["symbol"], row["strike"], row["opt_type"], row["expiry_date"], master)
            if key:
                df.at[idx, "instrument_key"] = key
                matched += 1
            elif status == "NOT FOUND":
                df.at[idx, "instrument_key"] = "NOT FOUND"
                notfound += 1
            else:
                df.at[idx, "instrument_key"] = status
                ambiguous += 1
        st.session_state.df = df
        st.success(f"Matched {matched} | Not found {notfound} | Ambiguous {ambiguous}")

open_keyed = df[(df["exit"].isin([None, ""])) & (~df["instrument_key"].isin(["", "NOT FOUND"]))
                 & (~df["instrument_key"].astype(str).str.startswith("AMBIGUOUS"))]

if col_c.button("Refresh Live LTP", type="primary"):
    if not st.session_state.access_token:
        st.error("Connect to Upstox first.")
    elif open_keyed.empty:
        st.warning("No open, keyed legs to refresh.")
    else:
        try:
            ltp_map = fetch_ltp(list(open_keyed["instrument_key"].unique()), st.session_state.access_token)
            now = datetime.now()
            for idx in open_keyed.index:
                price = ltp_map.get(df.at[idx, "instrument_key"])
                if price is not None:
                    df.at[idx, "ltp"] = price
                    df.at[idx, "last_updated"] = now
            st.session_state.df = df
            st.success(f"Refreshed {len(ltp_map)} instruments.")
        except Exception as e:
            st.error(str(e))

# ---------------------------------------------------------------- compute P&L

if "ltp" not in df.columns:
    df["ltp"] = None
if "last_updated" not in df.columns:
    df["last_updated"] = None


def leg_pnl(row):
    entry, exitp, qty, lot = row["entry"], row["exit"], row["qty"], row["lot_size"]
    if None in (entry, qty, lot):
        return None, None
    if exitp not in (None, ""):
        points = exitp - entry
        return points * qty, lot * points * qty
    if row["ltp"] is not None:
        points = row["ltp"] - entry
        return points * qty, lot * points * qty
    return None, None


df[["points", "profit"]] = df.apply(lambda r: pd.Series(leg_pnl(r)), axis=1)

st.subheader("2. Positions")
for sheet_name in df["sheet"].unique():
    sdf = df[df["sheet"] == sheet_name]
    st.markdown(f"**{sheet_name}**")
    agg = sdf.groupby("position_no").agg(
        symbol=("symbol", "first"), strategy=("strategy", "first"),
        entry_date=("entry_date", "first"),
        net_invest=("entry", lambda s: None),  # placeholder, recomputed below
    ).reset_index()

    # recompute net invest/profit per position properly
    rows_out = []
    for pos_no, g in sdf.groupby("position_no"):
        invest = sum((r["entry"] or 0) * (r["qty"] or 0) * (r["lot_size"] or 0) for _, r in g.iterrows())
        profit_sum = g["profit"].dropna().sum() if g["profit"].notna().any() else None
        pct = (profit_sum / invest * 100) if invest and profit_sum is not None else None
        rows_out.append({
            "Position": pos_no,
            "Symbol": g["symbol"].iloc[0],
            "Strategy": g["strategy"].iloc[0],
            "Legs": ", ".join(f"{r['strike_text']} @ {r['ltp'] if pd.notna(r['ltp']) else '-'}" for _, r in g.iterrows()),
            "Net Invest": round(invest, 2),
            "Live/Realized Net P&L": round(profit_sum, 2) if profit_sum is not None else None,
            "P&L %": round(pct, 2) if pct is not None else None,
            "Status": "closed" if g["exit"].notna().all() and (g["exit"] != "").all() else "open",
        })
    out_df = pd.DataFrame(rows_out)

    def style_pnl(v):
        if v is None or pd.isna(v):
            return ""
        return "color: #1a7f37" if v >= 0 else "color: #cf222e"

    styler = out_df.style
    styler = styler.map(style_pnl, subset=["Live/Realized Net P&L", "P&L %"]) if hasattr(styler, "map") \
        else styler.applymap(style_pnl, subset=["Live/Realized Net P&L", "P&L %"])
    st.dataframe(styler, width="stretch", hide_index=True)

    total_invest = out_df["Net Invest"].sum()
    total_pnl = out_df["Live/Realized Net P&L"].dropna().sum()
    c1, c2, c3 = st.columns(3)
    c1.metric("Net Invest", f"{total_invest:,.0f}")
    c2.metric("Net P&L", f"{total_pnl:,.0f}")
    c3.metric("P&L %", f"{(total_pnl/total_invest*100):.2f}%" if total_invest else "-")
    st.divider()

st.subheader("3. Export snapshot")
csv = df.to_csv(index=False).encode()
st.download_button("Download current data as CSV", csv, "option_pnl_snapshot.csv", "text/csv")

# regenerate positions_data.py so today's Expiry/Instrument Key fills survive a redeploy —
# download this and commit it over the existing positions_data.py in your repo.
export_cols = ["sheet", "position_no", "leg", "entry_date", "strategy", "symbol", "strike_text",
               "strike", "opt_type", "lot_size", "qty", "entry", "exit", "exit_date", "remarks",
               "expiry_date", "instrument_key"]
export_records = df[export_cols].copy()
for c in ("entry_date", "exit_date"):
    export_records[c] = export_records[c].apply(lambda v: v.isoformat() if pd.notna(v) else None)
export_records["expiry_date"] = export_records["expiry_date"].apply(lambda v: v.isoformat() if pd.notna(v) else None)
py_source = ('"""Bundled snapshot of Option_Strategy_PNL.xlsx, parsed into per-leg rows."""\n\nPOSITIONS = '
             + export_records.to_dict(orient="records").__repr__())
st.download_button("Download updated positions_data.py", py_source.encode(), "positions_data.py", "text/x-python")
